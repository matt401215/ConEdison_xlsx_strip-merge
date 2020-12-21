import os
import openpyxl
from pathlib import Path
from openpyxl import Workbook

###############################################################################           
#
# format_all_files()
# Cycles through all files in directory and creates a formatted copy of all
# xlsx files. Saves them in saveDirectory
#
###############################################################################
def format_all_files(directory,saveDirectory):    
    for filename in os.listdir(directory):
        if filename.endswith(".xlsx"):
            #print(filename)
            dataBase(os.path.join(directory,filename),saveDirectory,filename)
    print("Finished")


###############################################################################           
#
# dataBase()
# Reads in xlsx file located at filePath. Decided what format they are in and
# sends to the appropriate function to populate a new xlsx file in the
# universial format.
# The types are based on the number of columns the file has in its table(s).
# Below is the breakdown and order of columns by type.
#
#  Type 1            Type 2             Type 3             Type 4
#  date_posted       date_posted        date_posted        date_posted
#  days              days               days               days
#  bill_date         bill_date          bill_date          bill_date
#  elect_kwh_usage   elect_kwh_usage    elect_kwh_usage    elect_kwh_usage
#  elect_ce_charge   elect_ce_charge    elect_ce_charge    elect_ce_charge
#  gas_therm_usage   elect_esco         gas_therm_usage    elect_esco
#  gas_ce_charge     gas_therm_usage    gas_ce_charge      gas_therm_usage
#  tot_billing       gas_ce_charge      gas_esco           gas_ce_charge
#  other_charges     tot_billing        tot_billing        gas_esco 
#  payment           other_charges      other_charges      tot_billing
#  balance           payment            payment            other_charges 
#                    balance            balance            payment
#                                                          balance
#
###############################################################################
def dataBase(filePath,saveDirectory,currentFile):
    # Read xlsx file
    xlsx_obj = openpyxl.load_workbook(filePath)

    # Read the active sheet
    # Specify another sheet with xlsx_obj['Sheet_name']
    sheet = xlsx_obj.active

    # Determine the number of columns in the table by looking for the "balance"
    # column. sheet.max_column isn't working, returning every sheet as 12 even
    # when there are only 11 columns.
    # i: 0,1,2,3....rows-1
    # row: list of data (tuple)
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        # Cycle through all columns
        for k in range(sheet.max_column):
            if row[k] == "$\nBALANCE":
                tableWidth = k
    
    # Direct to correct template
    if tableWidth == 10:
        type1(sheet, tableWidth,saveDirectory,currentFile)
    elif tableWidth == 11:
        # Determine location of ESCO column
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            # Cycle through all columns
            for k in range(sheet.max_column):
                if row[k] == "ESCO\nSupply Charges":
                    escoLocation = k
        # Direct to appropriate function based on ESCO column location
        if escoLocation == 7:
            type3(sheet, tableWidth,saveDirectory,currentFile)
        elif escoLocation == 5:
            type2(sheet, tableWidth,saveDirectory,currentFile)
        else:
            print("Type 2/3 Error")
    elif tableWidth == 12:
        type4(sheet, tableWidth,saveDirectory,currentFile)
    else:
        print("column size error")

###############################################################################           
#
# get_address()
# 
#
###############################################################################
def get_address(sheet):
    count = 0
    for value in sheet.iter_rows(min_row=3, max_row=5, min_col=1,max_col=1,values_only=True):
        if count == 1:
            line1 = value
        elif count == 2:
            line2 = value
        else:
            line3 = value
        count += 1
    if line2[0] == "Dear Customer:":
        return line3, line1
    else:
        return line1, line2



###############################################################################           
#
# type1
# No ESCO columns
#
###############################################################################
def type1(sheet, tableWidth,saveDirectory,currentFile):
    date_posted = []
    days = []
    bill_date = []
    elect_kwh_usage = []
    elect_ce_charge =[]
    gas_therm_usage = []
    gas_ce_charge = []
    tot_billing = []
    other_charges = []
    payment = []
    balance = []       

    address1, address2 = get_address(sheet)
    #get_account_num(sheet)
    
    print("IN TYPE 1")
    tableStart = strip_find_top(sheet)

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i > tableStart:
            for k in range(tableWidth+1):
                #print('k = ', k, ' row[k] = ', row[k])
                if k == 0:
                    date_posted.append(row[k])
                elif k == 1:
                    days.append(row[k])
                elif k == 2:
                    bill_date.append(row[k])
                elif k == 3:
                    elect_kwh_usage.append(row[k])
                elif k == 4:
                    elect_ce_charge.append(row[k])
                elif k == 5:
                    gas_therm_usage.append(row[k])
                elif k == 6:
                    gas_ce_charge.append(row[k])
                elif k == 7:
                    tot_billing.append(row[k])
                elif k == 8:
                    other_charges.append(row[k])
                elif k == 9:
                    payment.append(row[k])
                elif k == 10:
                    balance.append(row[k])

    # Create new xlsx workbook and sheet
    NEWxlsx = Workbook()
    newSheet = NEWxlsx.active
    # Populate the headers
    newSheet["A1"] = 'address1'
    newSheet["B1"] = 'address2'
    newSheet["C1"] = 'date_posted'
    newSheet["D1"] = 'days'
    newSheet["E1"] = 'bill_date'
    newSheet["F1"] = 'elect_kwh_usage'
    newSheet["G1"] = 'elect_ce_charge'
    newSheet["H1"] = 'elect_esco'
    newSheet["I1"] = 'gas_therm_usage'
    newSheet["J1"] = 'gas_ce_charge'
    newSheet["K1"] = 'gas_esco'
    newSheet["L1"] = 'tot_billing'
    newSheet["M1"] = 'other_charges'
    newSheet["N1"] = 'balance'
    # Populate the correct columns with data    
    for i in range(len(date_posted)):
        newSheet["A"+ str(i+2)] = address1[0]
        newSheet["B"+ str(i+2)] = address2[0]
        newSheet["C"+ str(i+2)] = date_posted[i]
        newSheet["D"+ str(i+2)] = days[i]
        newSheet["E"+ str(i+2)] = bill_date[i]
        newSheet["F"+ str(i+2)] = elect_kwh_usage[i]
        newSheet["G"+ str(i+2)] = elect_ce_charge[i]
        #newSheet["H"+ str(i+2)] = elect_esco[i]
        newSheet["I"+ str(i+2)] = gas_therm_usage[i]
        newSheet["J"+ str(i+2)] = gas_ce_charge[i]
        #newSheet["K"+ str(i+2)] = gas_esco[i]
        newSheet["L"+ str(i+2)] = tot_billing[i]
        newSheet["M"+ str(i+2)] = other_charges[i]
        newSheet["N"+ str(i+2)] = balance[i]
    # Save new xlsx file
    NEWxlsx.save(Path(saveDirectory,"MOD_"+currentFile))


###############################################################################           
#
# type2
# Electric ESCO column only
#
###############################################################################
def type2(sheet, tableWidth,saveDirectory,currentFile):
    date_posted = []
    days = []
    bill_date = []
    elect_kwh_usage = []
    elect_ce_charge =[]
    elect_esco = []
    gas_therm_usage = []
    gas_ce_charge = []
    tot_billing = []
    other_charges = []
    payment = []
    balance = []

    address1, address2 = get_address(sheet)
    
    print("IN TYPE 2")
    tableStart = strip_find_top(sheet)
   
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i > tableStart:
            for k in range(tableWidth+1):
                #print('k = ', k, ' row[k] = ', row[k])
                if k == 0:
                    date_posted.append(row[k])
                elif k == 1:
                    days.append(row[k])
                elif k == 2:
                    bill_date.append(row[k])
                elif k == 3:
                    elect_kwh_usage.append(row[k])
                elif k == 4:
                    elect_ce_charge.append(row[k])
                elif k == 5:
                    elect_esco.append(row[k])
                elif k == 6:
                    gas_therm_usage.append(row[k])
                elif k == 7:
                    gas_ce_charge.append(row[k])
                elif k == 8:
                    tot_billing.append(row[k])
                elif k == 9:
                    other_charges.append(row[k])
                elif k == 10:
                    payment.append(row[k])
                elif k == 11:
                    balance.append(row[k])
    # Create new xlsx workbook and sheet
    NEWxlsx = Workbook()
    newSheet = NEWxlsx.active
    # Populate the headers
    newSheet["A1"] = 'address1'
    newSheet["B1"] = 'address2'
    newSheet["C1"] = 'date_posted'
    newSheet["D1"] = 'days'
    newSheet["E1"] = 'bill_date'
    newSheet["F1"] = 'elect_kwh_usage'
    newSheet["G1"] = 'elect_ce_charge'
    newSheet["H1"] = 'elect_esco'
    newSheet["I1"] = 'gas_therm_usage'
    newSheet["J1"] = 'gas_ce_charge'
    newSheet["K1"] = 'gas_esco'
    newSheet["L1"] = 'tot_billing'
    newSheet["M1"] = 'other_charges'
    newSheet["N1"] = 'balance'
    # Populate the correct columns with data
    for i in range(len(date_posted)):
        newSheet["A"+ str(i+2)] = address1[0]
        newSheet["B"+ str(i+2)] = address2[0]
        newSheet["C"+ str(i+2)] = date_posted[i]
        newSheet["D"+ str(i+2)] = days[i]
        newSheet["E"+ str(i+2)] = bill_date[i]
        newSheet["F"+ str(i+2)] = elect_kwh_usage[i]
        newSheet["G"+ str(i+2)] = elect_ce_charge[i]
        newSheet["H"+ str(i+2)] = elect_esco[i]
        newSheet["I"+ str(i+2)] = gas_therm_usage[i]
        newSheet["J"+ str(i+2)] = gas_ce_charge[i]
        #newSheet["K"+ str(i+2)] = gas_esco[i]
        newSheet["L"+ str(i+2)] = tot_billing[i]
        newSheet["M"+ str(i+2)] = other_charges[i]
        newSheet["N"+ str(i+2)] = balance[i]
    # Save new xlsx file
    NEWxlsx.save(Path(saveDirectory,"MOD_"+currentFile))


    
###############################################################################           
#
# type3
# Gas ESCO column only
#
###############################################################################
def type3(sheet, tableWidth,saveDirectory,currentFile):
    date_posted = []
    days = []
    bill_date = []
    elect_kwh_usage = []
    elect_ce_charge =[]
    gas_therm_usage = []
    gas_ce_charge = []
    gas_esco = []
    tot_billing = []
    other_charges = []
    payment = []
    balance = []

    address1, address2 = get_address(sheet)

    print("IN TYPE 3")
    tableStart = strip_find_top(sheet)
        
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i > tableStart:
            for k in range(tableWidth+1):
                if k == 0:
                    date_posted.append(row[k])
                elif k == 1:
                    days.append(row[k])
                elif k == 2:
                    bill_date.append(row[k])
                elif k == 3:
                    elect_kwh_usage.append(row[k])
                elif k == 4:
                    elect_ce_charge.append(row[k])
                elif k == 5:
                    gas_therm_usage.append(row[k])
                elif k == 6:
                    gas_ce_charge.append(row[k])
                elif k == 7:
                    gas_esco.append(row[k])
                elif k == 8:
                    tot_billing.append(row[k])
                elif k == 9:
                    other_charges.append(row[k])
                elif k == 10:
                    payment.append(row[k])
                elif k == 11:
                    balance.append(row[k])
    
    # Create new xlsx workbook and sheet
    NEWxlsx = Workbook()
    newSheet = NEWxlsx.active
    # Populate the headers
    newSheet["A1"] = 'address1'
    newSheet["B1"] = 'address2'
    newSheet["C1"] = 'date_posted'
    newSheet["D1"] = 'days'
    newSheet["E1"] = 'bill_date'
    newSheet["F1"] = 'elect_kwh_usage'
    newSheet["G1"] = 'elect_ce_charge'
    newSheet["H1"] = 'elect_esco'
    newSheet["I1"] = 'gas_therm_usage'
    newSheet["J1"] = 'gas_ce_charge'
    newSheet["K1"] = 'gas_esco'
    newSheet["L1"] = 'tot_billing'
    newSheet["M1"] = 'other_charges'
    newSheet["N1"] = 'balance'
    # Populate the correct columns with data
    for i in range(len(date_posted)):
        newSheet["A"+ str(i+2)] = address1[0]
        newSheet["B"+ str(i+2)] = address2[0]
        newSheet["C"+ str(i+2)] = date_posted[i]
        newSheet["D"+ str(i+2)] = days[i]
        newSheet["E"+ str(i+2)] = bill_date[i]
        newSheet["F"+ str(i+2)] = elect_kwh_usage[i]
        newSheet["G"+ str(i+2)] = elect_ce_charge[i]
        #newSheet["H"+ str(i+2)] = elect_esco[i]
        newSheet["I"+ str(i+2)] = gas_therm_usage[i]
        newSheet["J"+ str(i+2)] = gas_ce_charge[i]
        newSheet["K"+ str(i+2)] = gas_esco[i]
        newSheet["L"+ str(i+2)] = tot_billing[i]
        newSheet["M"+ str(i+2)] = other_charges[i]
        newSheet["N"+ str(i+2)] = balance[i]
    # Save new xlsx file
    NEWxlsx.save(Path(saveDirectory,"MOD_"+currentFile))

    
###############################################################################           
#
# type4
# Both ESCO columns
#
###############################################################################
def type4(sheet, tableWidth,saveDirectory,currentFile):
    date_posted = []
    days = []
    bill_date = []
    elect_kwh_usage = []
    elect_ce_charge =[]
    elect_esco = []
    gas_therm_usage = []
    gas_ce_charge = []
    gas_esco = []
    tot_billing = []
    other_charges = []
    payment = []
    balance = []

    address1, address2 = get_address(sheet)

    print("IN TYPE 4")
    tableStart = strip_find_top(sheet)

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i > tableStart:
            for k in range(tableWidth+1):
                if k == 0:
                    date_posted.append(row[k])
                elif k == 1:
                    days.append(row[k])
                elif k == 2:
                    bill_date.append(row[k])
                elif k == 3:
                    elect_kwh_usage.append(row[k])
                elif k == 4:
                    elect_ce_charge.append(row[k])
                elif k == 5:
                    elect_esco.append(row[k])
                elif k == 6:
                    gas_therm_usage.append(row[k])
                elif k == 7:
                    gas_ce_charge.append(row[k])
                elif k == 8:
                    gas_esco.append(row[k])
                elif k == 9:
                    tot_billing.append(row[k])
                elif k == 10:
                    other_charges.append(row[k])
                elif k == 11:
                    payment.append(row[k])
                elif k == 12:
                    balance.append(row[k])
                    
     # Create new xlsx workbook and sheet
    NEWxlsx = Workbook()
    newSheet = NEWxlsx.active
    # Populate the headers
    newSheet["A1"] = 'address1'
    newSheet["B1"] = 'address2'
    newSheet["C1"] = 'date_posted'
    newSheet["D1"] = 'days'
    newSheet["E1"] = 'bill_date'
    newSheet["F1"] = 'elect_kwh_usage'
    newSheet["G1"] = 'elect_ce_charge'
    newSheet["H1"] = 'elect_esco'
    newSheet["I1"] = 'gas_therm_usage'
    newSheet["J1"] = 'gas_ce_charge'
    newSheet["K1"] = 'gas_esco'
    newSheet["L1"] = 'tot_billing'
    newSheet["M1"] = 'other_charges'
    newSheet["N1"] = 'balance'
    # Populate the correct columns with data
    for i in range(len(date_posted)):
        newSheet["A"+ str(i+2)] = address1[0]
        newSheet["B"+ str(i+2)] = address2[0]
        newSheet["C"+ str(i+2)] = date_posted[i]
        newSheet["D"+ str(i+2)] = days[i]
        newSheet["E"+ str(i+2)] = bill_date[i]
        newSheet["F"+ str(i+2)] = elect_kwh_usage[i]
        newSheet["G"+ str(i+2)] = elect_ce_charge[i]
        newSheet["H"+ str(i+2)] = elect_esco[i]
        newSheet["I"+ str(i+2)] = gas_therm_usage[i]
        newSheet["J"+ str(i+2)] = gas_ce_charge[i]
        newSheet["K"+ str(i+2)] = gas_esco[i]
        newSheet["L"+ str(i+2)] = tot_billing[i]
        newSheet["M"+ str(i+2)] = other_charges[i]
        newSheet["N"+ str(i+2)] = balance[i]
    # Save new xlsx file
    NEWxlsx.save(Path(saveDirectory,"MOD_"+currentFile))


###############################################################################           
#
# strip_find_top()
# This function removes additional headers between tables so all data is in one
# table. It does this by counting how many headers there are, and then deleting
# all but the first. The function returns the row location of the first header
# Input: Sheet
# Output: int Location of first(only) header row
#
###############################################################################
def strip_find_top(sheet):
    tableS = []
    # Make a list of all the rows in which there is a table header
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        for k in range(sheet.max_column):
            if k == 0:
                if row[k] == "DATE POSTED":
                    tableS.append(i)
    # Delete the two header rows when more than one table is on a sheet
    if len(tableS) != 1:
        for ii in range(len(tableS)-1):
            rowNum = tableS[ii+1]
            sheet.delete_rows(rowNum)
            sheet.delete_rows(rowNum)

    return tableS[0]


    
###############################################################################           
#
# main()
# 
#
###############################################################################            
def main():

    #directory = '/Volumes/GoogleDrive/My Drive/Projects/Genesis Realty Group/Bill Auditing/Historical Utility and Supply Data, contracts/Con Edison/Excel Statement Summaries (Final)'
    #saveDirectory = '/Volumes/GoogleDrive/My Drive/Projects/Genesis Realty Group/Bill Auditing/Historical Utility and Supply Data, contracts/Con Edison/Excel Statement Summaries (Final)/Modified Final'
    directory = '/Users/matthewgilroy/Desktop/xlsx Test Folder'
    saveDirectory = '/Users/matthewgilroy/Desktop/xlsx Test Folder/AAA'

    format_all_files(directory,saveDirectory)
    


main()


